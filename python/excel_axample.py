import openpyxl
wb = openpyxl.load_workbook('D:\\f.xlsx')
sheet = wb.get_active_sheet()
for i in range(1, 358):
    sheet.cell(row=i, column=3).value = "insert into sendsms (smsindex,Phonenumber,Smscontent,Smstime,Smsuser,Status) values( SEQ_SENDSMS.nextval ," \
        + "\'" + str(sheet.cell(row=i, column=1).value) + "\'" + "," \
        + "'内容。'" \
        + ",sysdate,1,0)"

wb.save("D:\\f.xlsx")
